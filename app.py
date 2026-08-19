import os
import webbrowser
from threading import Timer
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from db import (
    init_db, get_elementos_ordenados, get_valores, set_valor,
    agregar_elemento, actualizar_elemento, eliminar_elemento,
    insertar_separador_despues_de, reordenar_elementos, obtener_elemento_por_id,
    get_periodos_activos, agregar_periodo_activo, eliminar_periodo_activo,
    kpi_activo_en_rango, get_valores_completos, importar_valores_masivos
)
from dates import obtener_estructura_periodos

app = Flask(__name__)


def calcular_valor_agregado(valores_diarios, calculo):
    """
    valores_diarios: lista de números (v_obj o v_real o v_gatillo)
    calculo: 'suma', 'promedio', 'maximo'
    Retorna el valor agregado o None.
    """
    if not valores_diarios:
        return None
    if calculo == 'suma':
        return sum(valores_diarios)
    elif calculo == 'promedio':
        return sum(valores_diarios) / len(valores_diarios)
    elif calculo == 'maximo':
        return max(valores_diarios)
    return None


def agregar_valores_periodo(kpi, periodo, valores_diarios_obj, valores_diarios_real, valores_diarios_gatillo):
    """
    Calcula obj, real, gatillo para un período (semanal/mensual) a partir de los diarios.
    """
    dias_del_periodo = periodo['dias']
    calculo = kpi['calculo']
    
    vals_obj = []
    vals_real = []
    vals_gatillo = []
    
    for dia in dias_del_periodo:
        dia_key = dia.date().isoformat()
        v_obj = valores_diarios_obj.get(dia_key)
        v_real = valores_diarios_real.get(dia_key)
        v_gat = valores_diarios_gatillo.get(dia_key)
        if v_obj is not None:
            vals_obj.append(v_obj)
        if v_real is not None:
            vals_real.append(v_real)
        if v_gat is not None:
            vals_gatillo.append(v_gat)
    
    obj_agregado = calcular_valor_agregado(vals_obj, calculo)
    real_agregado = calcular_valor_agregado(vals_real, calculo)
    gatillo_agregado = calcular_valor_agregado(vals_gatillo, calculo)
    
    return obj_agregado, real_agregado, gatillo_agregado


@app.route('/')
def index():
    return render_template('tablero.html')


@app.route('/maestro')
def maestro():
    return render_template('maestro.html')


@app.route('/api/tablero')
def api_tablero():
    try:
        # Obtener año actual si no se especifica
        year = int(request.args.get('year', datetime.now().year))
        tipo_reunion = request.args.get('tipo', 'diaria')
        periodos, meses_unicos = obtener_estructura_periodos(year, tipo_reunion)
        periodos_json = []
        for p in periodos:
            periodos_json.append({
                'nombre': p['nombre'],
                'nombre_sub': p.get('nombre_sub', ''),
                'fecha_inicio': p['fecha_inicio'].isoformat(),
                'fecha_fin': p['fecha_fin'].isoformat(),
                'dias': [d.isoformat() for d in p['dias']],
                'clase_mes': p['clase_mes']
            })
        
        elementos = get_elementos_ordenados()
        
        # Calcular rango de fechas visible en el tablero
        if periodos:
            rango_inicio = periodos[0]['fecha_inicio'].date().isoformat()
            rango_fin = periodos[-1]['fecha_fin'].date().isoformat()
        else:
            rango_inicio = rango_fin = f"{year}-01-01"

        # Filtrar KPIs por periodicidad y por períodos activos
        elementos_filtrados = []
        for elem in elementos:
            if elem['tipo'] == 'separador':
                elementos_filtrados.append(elem)
            elif elem['tipo'] == 'kpi':
                periodicidad_str = (elem.get('periodicidad') or '').strip().lower()
                if not periodicidad_str:
                    pass  # sin periodicidad definida no aplica filtro de reunión
                else:
                    periodicidades = [p.strip() for p in periodicidad_str.split(',') if p.strip()]
                    if tipo_reunion not in periodicidades:
                        continue
                if not kpi_activo_en_rango(elem['id'], rango_inicio, rango_fin):
                    continue
                elementos_filtrados.append(elem)
        
        kpis_filtrados = [e for e in elementos_filtrados if e['tipo'] == 'kpi']
        valores_obj = {}
        valores_real = {}
        valores_gatillo = {}
        
        for kpi in kpis_filtrados:
            obj_guardado, real_guardado, gatillo_guardado = get_valores(kpi['id'], year)
            obj_por_periodo = {}
            real_por_periodo = {}
            gatillo_por_periodo = {}
            
            if tipo_reunion == 'diaria':
                for p in periodos:
                    fecha_key = p['fecha_inicio'].date().isoformat()
                    obj_por_periodo[fecha_key] = obj_guardado.get(fecha_key, '')
                    real_por_periodo[fecha_key] = real_guardado.get(fecha_key, '')
                    gatillo_por_periodo[fecha_key] = gatillo_guardado.get(fecha_key, '')
            else:
                for p in periodos:
                    obj_calc, real_calc, gatillo_calc = agregar_valores_periodo(
                        kpi, p, obj_guardado, real_guardado, gatillo_guardado
                    )
                    fecha_key = p['fecha_inicio'].date().isoformat()
                    obj_por_periodo[fecha_key] = obj_calc if obj_calc is not None else ''
                    real_por_periodo[fecha_key] = real_calc if real_calc is not None else ''
                    gatillo_por_periodo[fecha_key] = gatillo_calc if gatillo_calc is not None else ''
            
            valores_obj[kpi['id']] = obj_por_periodo
            valores_real[kpi['id']] = real_por_periodo
            valores_gatillo[kpi['id']] = gatillo_por_periodo
        
        return jsonify({
            'periodos': periodos_json,
            'elementos': elementos_filtrados,
            'valoresObj': valores_obj,
            'valoresReal': valores_real,
            'valoresGatillo': valores_gatillo,
            'mesesUnicos': meses_unicos,
            'yearActual': datetime.now().year,
            'mesActual': datetime.now().month
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/valor', methods=['POST'])
def save_valor():
    data = request.json
    set_valor(data['kpi_id'], data['fecha'], data['tipo'], data['valor'])
    return jsonify({'ok': True})


@app.route('/api/elementos', methods=['GET'])
def api_list_elementos():
    elementos = get_elementos_ordenados()
    return jsonify(elementos)


@app.route('/api/elemento/<int:elemento_id>', methods=['GET'])
def api_get_elemento(elemento_id):
    elem = obtener_elemento_por_id(elemento_id)
    return jsonify(elem)


@app.route('/api/elemento', methods=['POST'])
def api_create_elemento():
    data = request.json
    try:
        new_id = agregar_elemento(data)
        return jsonify({'id': new_id})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/elemento/<int:elemento_id>', methods=['PUT'])
def api_update_elemento(elemento_id):
    data = request.json
    try:
        actualizar_elemento(elemento_id, data)
        return jsonify({'ok': True})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/elemento/<int:elemento_id>', methods=['DELETE'])
def api_delete_elemento(elemento_id):
    eliminar_elemento(elemento_id)
    return jsonify({'ok': True})


@app.route('/api/separador', methods=['POST'])
def api_insertar_separador():
    data = request.json
    elemento_id = data.get('after_id')
    titulo = data.get('titulo', '---')
    if elemento_id:
        insertar_separador_despues_de(elemento_id, titulo)
    else:
        from db import agregar_elemento
        agregar_elemento({
            'tipo': 'separador',
            'descripcion': titulo,
            'codigo': None, 'duenio': '', 'unidad': '', 'calculo': '',
            'polaridad': '', 'definicion': '', 'objetivo': '', 'forma_calculo': '',
            'excluye': '', 'periodicidad': '', 'fuente_info': '', 'es_critico': 'no'
        })
    return jsonify({'ok': True})


@app.route('/api/reordenar', methods=['POST'])
def api_reordenar():
    ordenes = request.json.get('ordenes', [])
    reordenar_elementos(ordenes)
    return jsonify({'ok': True})


@app.route('/api/elemento/<int:kpi_id>/periodos', methods=['GET'])
def api_get_periodos(kpi_id):
    return jsonify(get_periodos_activos(kpi_id))


@app.route('/api/elemento/<int:kpi_id>/periodos', methods=['POST'])
def api_agregar_periodo(kpi_id):
    data = request.json
    new_id = agregar_periodo_activo(kpi_id, data['fecha_inicio'], data['fecha_fin'])
    return jsonify({'id': new_id})


@app.route('/api/periodo/<int:periodo_id>', methods=['DELETE'])
def api_eliminar_periodo(periodo_id):
    eliminar_periodo_activo(periodo_id)
    return jsonify({'ok': True})


# ── FUNCIONES DE EXPORTACIÓN E IMPORTACIÓN CON XLSX ─────────────────────────

def estilo_excel(wb, ws, titulo=None):
    """Aplica estilos básicos a una hoja de Excel"""
    try:
        # Fuente para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Borde fino
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar a la primera fila (encabezados)
        if ws.max_row > 0 and ws.max_column > 0:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        # Aplicar bordes a todas las celdas
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.border:
                    cell.border = thin_border
                if row > 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = ws.column_dimensions[openpyxl.utils.get_column_letter(col)]
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            column.width = min(max_length + 2, 40)
        
        # Si hay título, agregarlo
        if titulo and ws.max_row > 0:
            ws.insert_rows(0)
            ws.cell(row=1, column=1, value=titulo)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    except Exception as e:
        # Si hay error en estilos, continuamos sin estilos
        print(f"Error al aplicar estilos: {e}")


@app.route('/api/exportar/maestro', methods=['GET'])
def exportar_maestro():
    """Exporta todos los elementos del maestro a XLSX"""
    try:
        elementos = get_elementos_ordenados()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maestro KPI"
        
        # Cabeceras
        headers = [
            'id', 'tipo', 'codigo', 'descripcion', 'duenio', 'unidad', 'calculo',
            'polaridad', 'definicion', 'objetivo', 'forma_calculo', 'excluye',
            'periodicidad', 'fuente_info', 'es_critico', 'orden'
        ]
        ws.append(headers)
        
        # Datos
        for elem in elementos:
            ws.append([
                elem.get('id', ''),
                elem.get('tipo', ''),
                elem.get('codigo') or '',
                elem.get('descripcion') or '',
                elem.get('duenio') or '',
                elem.get('unidad') or '',
                elem.get('calculo') or '',
                elem.get('polaridad') or '',
                elem.get('definicion') or '',
                elem.get('objetivo') or '',
                elem.get('forma_calculo') or '',
                elem.get('excluye') or '',
                elem.get('periodicidad') or '',
                elem.get('fuente_info') or '',
                elem.get('es_critico') or 'no',
                elem.get('orden') or 0
            ])
        
        estilo_excel(wb, ws, "MAESTRO DE KPI - Exportación")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='maestro_kpi.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/exportar/valores', methods=['GET'])
def exportar_valores():
    """Exporta todos los valores a XLSX con formato profesional"""
    try:
        kpis = get_elementos_ordenados()
        kpis_kpi = [k for k in kpis if k['tipo'] == 'kpi']
        
        wb = openpyxl.Workbook()
        
        # Hoja 1: Valores planos (para importación)
        ws_flat = wb.active
        ws_flat.title = "Valores"
        ws_flat.append(['kpi_id', 'codigo', 'descripcion', 'fecha', 'tipo', 'valor'])
        
        # Hoja 2: Resumen por KPI
        ws_summary = wb.create_sheet("Resumen por KPI")
        ws_summary.append(['Código', 'Descripción', 'Dueño', 'Unidad', 'Cálculo', 'Polaridad', 'Total Obj', 'Total Real', 'Total Gatillo'])
        
        total_importados = 0
        
        for kpi in kpis_kpi:
            try:
                # Obtener valores usando get_valores_completos (más eficiente)
                obj_data, real_data, gatillo_data = get_valores_completos(kpi['id'])
                
                # Si get_valores_completos no devuelve datos, intentar con get_valores para años específicos
                if not obj_data and not real_data and not gatillo_data:
                    for year in range(2020, 2031):
                        obj, real, gatillo = get_valores(kpi['id'], year)
                        obj_data.update(obj)
                        real_data.update(real)
                        gatillo_data.update(gatillo)
                
                # Escribir valores planos
                for fecha, valor in obj_data.items():
                    if valor is not None and valor != '':
                        ws_flat.append([kpi['id'], kpi.get('codigo') or '', kpi.get('descripcion') or '', fecha, 'obj', float(valor) if valor else ''])
                        total_importados += 1
                
                for fecha, valor in real_data.items():
                    if valor is not None and valor != '':
                        ws_flat.append([kpi['id'], kpi.get('codigo') or '', kpi.get('descripcion') or '', fecha, 'real', float(valor) if valor else ''])
                        total_importados += 1
                
                for fecha, valor in gatillo_data.items():
                    if valor is not None and valor != '':
                        ws_flat.append([kpi['id'], kpi.get('codigo') or '', kpi.get('descripcion') or '', fecha, 'gatillo', float(valor) if valor else ''])
                        total_importados += 1
                
                # Resumen por KPI
                total_obj = sum(float(v) for v in obj_data.values() if v is not None and v != '') if obj_data else 0
                total_real = sum(float(v) for v in real_data.values() if v is not None and v != '') if real_data else 0
                total_gatillo = sum(float(v) for v in gatillo_data.values() if v is not None and v != '') if gatillo_data else 0
                
                ws_summary.append([
                    kpi.get('codigo') or '',
                    kpi.get('descripcion') or '',
                    kpi.get('duenio') or '',
                    kpi.get('unidad') or '',
                    kpi.get('calculo') or '',
                    kpi.get('polaridad') or '',
                    total_obj,
                    total_real,
                    total_gatillo
                ])
                
            except Exception as e:
                # Si falla un KPI, continuamos con el siguiente
                print(f"Error procesando KPI {kpi.get('id')}: {e}")
                continue
        
        # Aplicar estilos a ambas hojas
        estilo_excel(wb, ws_flat, f"VALORES - Total: {total_importados} registros")
        estilo_excel(wb, ws_summary, "RESUMEN POR KPI")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='valores_kpi.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/importar/maestro', methods=['POST'])
def importar_maestro():
    """Importa elementos del maestro desde XLSX"""
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400
    
    try:
        # Leer archivo XLSX
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        
        # Obtener encabezados (primera fila)
        headers = [cell.value for cell in ws[1]]
        
        elementos_importados = []
        errores = []
        
        # Obtener elementos existentes para evitar duplicados
        existentes = get_elementos_ordenados()
        codigos_existentes = [e['codigo'] for e in existentes if e['tipo'] == 'kpi' and e['codigo']]
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
                
            if len(row) < 16:
                errores.append(f"Fila {row_num}: Número de columnas incorrecto ({len(row)} esperado 16)")
                continue
            
            try:
                elem_data = {
                    'tipo': str(row[1] or '').strip() or 'kpi',
                    'codigo': str(row[2] or '').strip() or None,
                    'descripcion': str(row[3] or '').strip() or '',
                    'duenio': str(row[4] or '').strip() or '',
                    'unidad': str(row[5] or '').strip() or '',
                    'calculo': str(row[6] or '').strip() or '',
                    'polaridad': str(row[7] or '').strip() or '',
                    'definicion': str(row[8] or '').strip() or '',
                    'objetivo': str(row[9] or '').strip() or '',
                    'forma_calculo': str(row[10] or '').strip() or '',
                    'excluye': str(row[11] or '').strip() or '',
                    'periodicidad': str(row[12] or '').strip() or '',
                    'fuente_info': str(row[13] or '').strip() or '',
                    'es_critico': str(row[14] or '').strip() or 'no',
                }
                
                # Validar
                if elem_data['tipo'] == 'kpi':
                    if not elem_data['codigo']:
                        errores.append(f"Fila {row_num}: El código es obligatorio para KPIs")
                        continue
                    if elem_data['codigo'] in codigos_existentes:
                        errores.append(f"Fila {row_num}: El código '{elem_data['codigo']}' ya existe")
                        continue
                    if not elem_data['descripcion']:
                        errores.append(f"Fila {row_num}: La descripción es obligatoria")
                        continue
                
                # Agregar el elemento
                new_id = agregar_elemento(elem_data)
                elementos_importados.append(new_id)
                
            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")
        
        return jsonify({
            'ok': True,
            'importados': len(elementos_importados),
            'errores': errores
        })
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 400


@app.route('/api/importar/valores', methods=['POST'])
def importar_valores():
    """Importa valores desde XLSX"""
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400
    
    try:
        # Leer archivo XLSX
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        
        # Obtener mapeo de códigos a IDs
        elementos = get_elementos_ordenados()
        codigo_to_id = {e['codigo']: e['id'] for e in elementos if e['tipo'] == 'kpi' and e['codigo']}
        
        valores_importados = 0
        errores = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
                
            if len(row) < 6:
                errores.append(f"Fila {row_num}: Número de columnas incorrecto ({len(row)} esperado 6)")
                continue
            
            try:
                # Formato: kpi_id, codigo, descripcion, fecha, tipo, valor
                codigo = str(row[1] or '').strip()
                fecha = str(row[3] or '').strip()
                tipo = str(row[4] or '').strip().lower()
                valor_str = str(row[5] or '').strip()
                
                if not codigo or codigo not in codigo_to_id:
                    errores.append(f"Fila {row_num}: Código '{codigo}' no encontrado")
                    continue
                
                if not fecha:
                    errores.append(f"Fila {row_num}: Fecha vacía")
                    continue
                
                if tipo not in ['obj', 'real', 'gatillo']:
                    errores.append(f"Fila {row_num}: Tipo inválido '{tipo}' (debe ser obj, real o gatillo)")
                    continue
                
                try:
                    valor = float(valor_str) if valor_str else None
                except ValueError:
                    errores.append(f"Fila {row_num}: Valor '{valor_str}' no es numérico")
                    continue
                
                if valor is not None:
                    kpi_id = codigo_to_id[codigo]
                    set_valor(kpi_id, fecha, tipo, valor)
                    valores_importados += 1
                
            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")
        
        return jsonify({
            'ok': True,
            'importados': valores_importados,
            'errores': errores
        })
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 400


@app.route('/api/plantilla/maestro', methods=['GET'])
def plantilla_maestro():
    """Descarga plantilla para maestro KPI en XLSX"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Maestro KPI"
    
    # Encabezados
    headers = [
        'id', 'tipo', 'codigo', 'descripcion', 'duenio', 'unidad', 'calculo',
        'polaridad', 'definicion', 'objetivo', 'forma_calculo', 'excluye',
        'periodicidad', 'fuente_info', 'es_critico', 'orden'
    ]
    ws.append(headers)
    
    # Ejemplo de KPI
    ws.append([
        '', 'kpi', '1.1', 'Lesiones con tarea modificada', 'Ref Seguridad', 
        '#', 'suma', '▼', 'Número de lesiones que modifican tarea', 
        'Reducir a 0', 'Suma de eventos', 'Ninguna', 'diaria', 'Registro interno', 'no', ''
    ])
    
    # Ejemplo de separador
    ws.append([
        '', 'separador', '', 'SEGURIDAD', '', '', '', '', '', '', '', '', '', '', '', ''
    ])
    
    # Ejemplo de otro KPI
    ws.append([
        '', 'kpi', '1.2', 'Días sin accidentes', 'Ref Seguridad',
        '#', 'suma', '▲', 'Días consecutivos sin accidentes',
        'Mantener tendencia', 'Conteo diario', 'Ninguna', 'diaria', 'Registro interno', 'no', ''
    ])
    
    estilo_excel(wb, ws, "PLANTILLA MAESTRO KPI - Complete los datos")
    
    # Agregar hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    ws_inst.append(["INSTRUCCIONES PARA COMPLETAR LA PLANTILLA MAESTRO KPI"])
    ws_inst.append([""])
    ws_inst.append(["1. Complete los datos en la hoja 'Plantilla Maestro KPI'"])
    ws_inst.append(["2. No modifique la estructura de columnas (16 columnas)"])
    ws_inst.append(["3. Para KPIs, el campo 'codigo' es obligatorio y debe ser único"])
    ws_inst.append(["4. Para separadores, solo complete 'tipo' (separador) y 'descripcion'"])
    ws_inst.append(["5. Periodicidad: diaria, semanal, mensual (puede combinar con comas)"])
    ws_inst.append(["6. Es_critico: 'si' o 'no'"])
    ws_inst.append(["7. Los campos marcados con * son obligatorios para KPIs"])
    ws_inst.append([""])
    ws_inst.append(["Campos obligatorios (*): codigo, descripcion, duenio, unidad, calculo,"])
    ws_inst.append(["polaridad, definicion, objetivo, forma_calculo, excluye, periodicidad,"])
    ws_inst.append(["fuente_info"])
    
    # Estilo para instrucciones
    for row in ws_inst.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_inst.column_dimensions['A'].width = 60
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_maestro_kpi.xlsx'
    )


@app.route('/api/plantilla/valores', methods=['GET'])
def plantilla_valores():
    """Descarga plantilla para valores KPI en XLSX"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Valores"
    
    # Encabezados
    ws.append(['kpi_id', 'codigo', 'descripcion', 'fecha', 'tipo', 'valor'])
    
    # Ejemplos de valores
    ws.append(['', '1.1', 'Lesiones con tarea modificada', '2026-01-01', 'obj', '0'])
    ws.append(['', '1.1', 'Lesiones con tarea modificada', '2026-01-01', 'real', '0'])
    ws.append(['', '1.1', 'Lesiones con tarea modificada', '2026-01-01', 'gatillo', '3'])
    ws.append(['', '1.1', 'Lesiones con tarea modificada', '2026-01-02', 'obj', '0'])
    ws.append(['', '1.1', 'Lesiones con tarea modificada', '2026-01-02', 'real', '1'])
    ws.append(['', '1.2', 'Días sin accidentes', '2026-01-01', 'obj', '0'])
    ws.append(['', '1.2', 'Días sin accidentes', '2026-01-01', 'real', '5'])
    
    estilo_excel(wb, ws, "PLANTILLA VALORES KPI - Complete los datos")
    
    # Agregar hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    ws_inst.append(["INSTRUCCIONES PARA COMPLETAR LA PLANTILLA VALORES KPI"])
    ws_inst.append([""])
    ws_inst.append(["1. Complete los datos en la hoja 'Plantilla Valores'"])
    ws_inst.append(["2. El campo 'codigo' debe coincidir con el código de un KPI existente"])
    ws_inst.append(["3. La 'descripcion' es opcional y solo para referencia"])
    ws_inst.append(["4. El campo 'fecha' debe tener formato YYYY-MM-DD"])
    ws_inst.append(["5. El campo 'tipo' debe ser: obj, real o gatillo"])
    ws_inst.append(["6. El campo 'valor' debe ser numérico"])
    ws_inst.append(["7. Si el valor está vacío, se ignorará la fila"])
    ws_inst.append([""])
    ws_inst.append(["Tipo de valores:"])
    ws_inst.append(["  - obj: Valor objetivo (meta)"])
    ws_inst.append(["  - real: Valor real (medición)"])
    ws_inst.append(["  - gatillo: Valor de gatillo (solo para KPIs críticos)"])
    
    for row in ws_inst.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_inst.column_dimensions['A'].width = 60
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_valores_kpi.xlsx'
    )


if __name__ == '__main__':
    # Inicializar la base de datos
    init_db()
    
    # Obtener el puerto desde la variable de entorno (Render asigna uno)
    port = int(os.environ.get('PORT', 5000))
    
    # Detectar si estamos en producción o desarrollo
    is_production = os.environ.get('RENDER', False)
    
    if not is_production:
        # Solo abrir navegador en desarrollo local
        def abrir_navegador():
            webbrowser.open_new('http://127.0.0.1:5000')
        Timer(1, abrir_navegador).start()
    
    # Ej
